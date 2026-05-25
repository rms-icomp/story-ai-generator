from rest_framework.response import Response
from rest_framework.decorators import api_view
import requests
from django.shortcuts import render
from django.db import IntegrityError


from app.models import HistoriaUsuario, Produto, Usuario
from .forms import InformacaoForm
from .prompts import prepararPrompt
import re
import json
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font

def exportar_historias_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Apresentação do Produto e Historias"

    bold = Font(bold=True)

    # Cabeçalho principal
    ws.append(["Produto", "Usuário / História", "Descrição", "Ações / Detalhes"])
    for cell in ws[1]:
        cell.font = bold

    for produto in Produto.objects.all():
        # ----------------------------------------
        # Linha do Produto
        # ----------------------------------------
        ws.append([
            f"Produto #{produto.id}",
            "",
            produto.apresentacao_produto,
            produto.informacoes_adicionais or ""
        ])

        # ----------------------------------------
        # Usuários relacionados
        # ----------------------------------------
        usuarios = produto.usuarios.all()

        if usuarios:
            ws.append(["", "Usuários do Produto", "", ""])
            ws.append(["", "Função / Papel", "Descrição", "Ações"])
            for cell in ws[ws.max_row]:
                cell.font = bold

            for u in usuarios:
                ws.append([
                    "",
                    u.usuario_papel,
                    u.descricao_usuario,
                    u.acoes_usuario
                ])
        else:
            ws.append(["", "Nenhum usuário cadastrado", "", ""])

        ws.append([""])  # linha em branco

        # ----------------------------------------
        # Histórias relacionadas
        # ----------------------------------------
        historias = produto.historias.all()

        if historias:
            ws.append(["", "Histórias do Produto", "", ""])
            ws.append(["", "Título", "Como", "Eu quero / Para que"])
            for cell in ws[ws.max_row]:
                cell.font = bold

            for h in historias:
                ws.append([
                    "",
                    h.titulo,
                    f"Como {h.como}",
                    f"Eu quero {h.eu_quero}, para que {h.para_que}"
                ])
        else:
            ws.append(["", "Nenhuma história cadastrada", "", ""])

        ws.append([""])  # espaço entre produtos
        ws.append([""])

    # ----------------------------------------
    # Resultado
    # ----------------------------------------
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="produtos_completos.xlsx"'

    wb.save(response)
    return response

@api_view(['POST'])  # Aceita apenas requisições POST
def executar_prompt(request):

    print("executar_prompt")

    """Recebe os dados e retorna uma resposta formatada
    apresentacao   = request.data.get('apresentacao')
    usuarios       = request.data.get('usuarios')
    requisitos     = request.data.get('requisitos')"""

    resultado = prepararPrompt(request.data)

    print("executar_prompt > resultado ")
    print(resultado.choices[0].message.content)

    
    try:
        conteudo = resultado.choices[0].message.content
    except Exception as e:
        return Response({"erro": "Erro ao acessar conteúdo da resposta", "detalhe": str(e)}, status=500)

    return Response(conteudo, status=201)

def executar_prompt_test(dados_):

    resultado = prepararPrompt(dados_)

    print("executar_prompt > resultado ")

    try:
        #conteudo = resultado.choices[0].message.content
        conteudo = resultado
    except Exception as e:
        return Response({"erro": "Erro ao acessar conteúdo da resposta", "detalhe": str(e)}, status=500)

    return conteudo

def getQtdeStory(texto):
    print("getQtdeStory")
    
    # Regex que captura cada história individualmente
    padrao_historia = re.compile(
        r"(Hist[óo]ria de Usu[áa]rio(?: Adicional)?\s+\d+:.*?)(?=(\nHist[óo]ria de Usu[áa]rio)|$)",
        re.DOTALL | re.IGNORECASE
    )

    # Extrair todas as histórias
    historias = []

    for historia_bloco in padrao_historia.findall(texto):
        bloco = historia_bloco[0]

        # Regex para separar título, persona, desejo e objetivo dentro do bloco
        detalhe = re.search(
            r"Hist[óo]ria de Usu[áa]rio(?: Adicional)?\s+(\d+):\s*(.*?)\n\s*- Como (.*?),\n\s*Eu quero (.*?),\n\s*Para que (?:eu )?(.*?)(?:\.|\n)",
            bloco, re.DOTALL | re.IGNORECASE
        )

        if detalhe:
            id_historia, titulo, persona, desejo, objetivo = detalhe.groups()
            historias.append({
                "id": int(id_historia),
                "titulo": titulo.strip(),
                "como": persona.strip(),
                "eu_quero": desejo.strip(),
                "para_que": objetivo.strip()
            })

    # Gerar JSON final
    resultado = {
        "historias": historias
    }

    print(resultado)

    return resultado

def salvarDadosBanco(dadosForm, historiasJson):
    
    try:
        produto = Produto.objects.create(
            apresentacao_produto = dadosForm['apresentacao'],
            informacoes_adicionais = dadosForm['requisitos']
        )

        for i in range(len(dadosForm['usuarios'])):
            usuario = Usuario.objects.create(
                produto = produto,
                usuario_papel= dadosForm['usuarios'][i][0] ,
                descricao_usuario= dadosForm['usuarios'][i][1] ,
                acoes_usuario= dadosForm['usuarios'][i][2] 
            )

        for h in historiasJson["historias"]:
            #print(h["titulo"], h["como"], h["eu_quero"], h["para_que"])

            historia = HistoriaUsuario.objects.create(
                produto=produto,
                titulo=h["titulo"],
                como= h["como"],
                eu_quero= h["eu_quero"],
                para_que= h["para_que"]
            )
        print("Salvo no banco com sucesso!")

    except IntegrityError as e:
        print("Erro ao salvar no banco:", e)


def formulario_view(request):
    mensagem = None

    if request.method == 'POST':

        tam = len(request.POST.getlist("user_nome"))
        
        usuarios = []
        for i in range(tam):
            nome      = request.POST.getlist("user_nome")[i]
            descricao = request.POST.getlist("user_descricao")[i]
            acoes     = request.POST.getlist("user_acoes")[i]
            usuarios.append([nome, descricao, acoes])

        
        dados = {
            "apresentacao": request.POST.get('apresentacao'),
            "usuarios": usuarios,
            "requisitos": request.POST.get('requisitos')
        }

        print("form")
        print(usuarios)

        # Enviando os dados via POST para a API
        #url = "http://localhost:8000/executar-prompt/"
        #response = requests.post(url, json=dados)

        response = executar_prompt_test(dados)

        print("response")

        #mensagem = getQtdeStory(response)

        """if response.status_code == 201:
            mensagem = response.text
            #mensagem = mensagem.replace("\\n", "<br>")
        else:
            mensagem = f"Erro ao enviar: {response.json()}"""
        # Converte string JSON em dicionário Python
        dados_json = json.loads(response)

        salvarDadosBanco(dados, dados_json)

        print(dados_json)

        return render(request, 'api/form-stories.html', {'mensagem': dados_json})    

    else:
        form = InformacaoForm()

        return render(request, 'api/formulario.html', {'form': form, 'mensagem': mensagem})